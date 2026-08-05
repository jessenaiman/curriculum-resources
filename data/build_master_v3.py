#!/usr/bin/env python3
"""
Build Master_Song_Curriculum_Sheet.xlsx — v3 with proper category grouping.
6 sheets: Master Song List, Daycare (0-2), Preschool (3-4), Kindergarten (5), Grade 1, Grade 2
"""

import sqlite3
import json
from collections import defaultdict, OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = r"C:\Users\jesse\OneDrive\Documents\New project\data\curriculum.db"
TAGS_PATH = r"C:\Users\jesse\OneDrive\Documents\New project\data\song_tags_clean.json"
OUTPUT = r"C:\Users\jesse\OneDrive\Documents\New project\data\Master_Song_Curriculum_Sheet.xlsx"

# === STYLES ===
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
CAT_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
CAT_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUB_FONT = Font(name="Calibri", bold=True, size=10, color="1F3864")
SUB_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOPIC_FONT = Font(name="Calibri", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

def infer_type(name, actions, topic):
    c = ((name or "") + " " + (actions or "") + " " + (topic or "")).lower()
    if any(w in c for w in ["fingerplay", "finger play", "fingers"]): return "Fingerplay"
    if any(w in c for w in ["bounce", "bouncing", "lap"]): return "Bounce"
    if any(w in c for w in ["movement", "march", "dance", "shake", "wiggle", "jump", "hop"]): return "Movement"
    if any(w in c for w in ["lullaby", "sleep", "night", "moon", "star"]): return "Lullaby"
    if any(w in c for w in ["game", "tag", "hide", "chase"]): return "Game"
    if any(w in c for w in ["transition", "clean up", "lining up", "wash"]): return "Transition"
    if any(w in c for w in ["flannel", "felt", "board"]): return "Flannel Board"
    return "Song"

def infer_materials(actions):
    if not actions: return ""
    a = actions.lower()
    m = []
    if any(w in a for w in ["scarf", "silk"]): m.append("Scarves")
    if any(w in a for w in ["instrument", "drum", "shaker", "rattle", "bell", "tambourine"]): m.append("Instruments")
    if any(w in a for w in ["puppet", "stuffed"]): m.append("Puppets")
    if any(w in a for w in ["flannel", "felt", "board"]): m.append("Flannel Board")
    if any(w in a for w in ["ball", "beanbag"]): m.append("Ball/Beanbag")
    if any(w in a for w in ["crayon", "paint", "paper", "glue"]): m.append("Art Supplies")
    if not m: m.append("Hands")
    return ", ".join(m)

def write_header(ws, row_num, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row_num, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = THIN
    ws.auto_filter.ref = f"A{row_num}:{get_column_letter(len(headers))}{row_num}"
    ws.freeze_panes = f"A{row_num + 1}"

def write_row(ws, rn, vals, font=None, fill=None):
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=rn, column=ci, value=v)
        if font: c.font = font
        if fill: c.fill = fill
        c.alignment = WRAP
        c.border = THIN

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# LOAD DATA
# ============================================================
conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row

songs = {r["id"]: dict(r) for r in conn.execute("SELECT * FROM songs").fetchall()}
ct_songs = {r["id"]: dict(r) for r in conn.execute("SELECT * FROM circle_time_songs").fetchall()}

g1_topics = [dict(r) for r in conn.execute(
    "SELECT * FROM curriculum_topics WHERE grade='Grade 1' ORDER BY subject, category, seq_num").fetchall()]
g2_topics = [dict(r) for r in conn.execute(
    "SELECT * FROM curriculum_topics WHERE grade='Grade 2' ORDER BY subject, category, seq_num").fetchall()]
g12_topics = [dict(r) for r in conn.execute(
    "SELECT * FROM curriculum_topics WHERE grade='Grade 1-2' ORDER BY subject, category, seq_num").fetchall()]

eyt_all = [dict(r) for r in conn.execute("SELECT * FROM early_years_topics ORDER BY id").fetchall()]

# Assign categories to early years data rows by cascading from headers
current_cat = None
current_sub = None
for t in eyt_all:
    seq = t.get("seq_num")
    cat_val = t.get("category")
    sub_val = t.get("subcategory")
    if seq is None or seq == "":
        # Header row — update running state
        if cat_val:
            current_cat = cat_val
        if sub_val:
            current_sub = sub_val
        elif cat_val is None:
            pass  # meta note row, keep current
    else:
        # Data row — inherit category
        if cat_val:
            current_cat = cat_val
        if sub_val:
            current_sub = sub_val
        t["_category"] = current_cat or "(Uncategorized)"
        t["_subcategory"] = current_sub or ""

# Join tables
song_curriculum = defaultdict(list)
for r in conn.execute("SELECT * FROM songs_curriculum").fetchall():
    song_curriculum[r["song_id"]].append({"curriculum_id": r["curriculum_id"], "relevance": r["relevance"]})

song_early_years = defaultdict(list)
for r in conn.execute("SELECT * FROM songs_early_years").fetchall():
    song_early_years[r["song_id"]].append({"early_years_id": r["early_years_id"], "relevance": r["relevance"]})

ct_curriculum = defaultdict(list)
for r in conn.execute("SELECT * FROM circle_time_songs_curriculum").fetchall():
    ct_curriculum[r["circle_time_song_id"]].append({"curriculum_id": r["curriculum_topic_id"], "relevance": r["relevance"]})

ct_early_years = defaultdict(list)
for r in conn.execute("SELECT * FROM circle_time_songs_early_years").fetchall():
    ct_early_years[r["circle_time_song_id"]].append({"early_years_id": r["early_years_topic_id"], "relevance": r["relevance"]})

# Lookups
curriculum_by_id = {}
for topics in [g1_topics, g2_topics, g12_topics]:
    for t in topics:
        curriculum_by_id[t["id"]] = t

eyt_by_id = {t["id"]: t for t in eyt_all}

with open(TAGS_PATH, "r", encoding="utf-8") as f:
    song_tags = json.load(f)

conn.close()

# Build reverse lookups: curriculum_id → song names
curr_to_songs = defaultdict(list)
for sid, links in song_curriculum.items():
    name = songs.get(sid, {}).get("song_name", "")
    for lk in links:
        curr_to_songs[lk["curriculum_id"]].append(name)

# early_years_id → song names
ey_to_songs = defaultdict(list)
for sid, links in song_early_years.items():
    name = songs.get(sid, {}).get("song_name", "")
    for lk in links:
        ey_to_songs[lk["early_years_id"]].append(name)

print(f"Loaded: {len(songs)} songs, {len(ct_songs)} circle time")
print(f"Curriculum: G1={len(g1_topics)}, G2={len(g2_topics)}, G1-2={len(g12_topics)}")
print(f"Curriculum → songs: {len(curr_to_songs)} topics linked")
print(f"Early years → songs: {len(ey_to_songs)} topics linked")


# ============================================================
# CREATE WORKBOOK
# ============================================================
wb = Workbook()
wb.remove(wb.active)


# ============================================================
# SHEET 1: MASTER SONG LIST
# ============================================================
ws1 = wb.create_sheet("Master Song List")
h1 = ["Song Name", "Source", "Age Range", "Type", "Educational Domain",
      "Tags", "Lyrics", "Actions", "Materials Needed", "Creator/Artist",
      "Source Title", "Curriculum Links", "Early Years Links"]
write_header(ws1, 1, h1)
set_widths(ws1, [30, 22, 18, 14, 22, 40, 50, 40, 20, 20, 25, 50, 50])

row = 2
for sid in sorted(songs.keys()):
    s = songs[sid]
    tags_str = song_tags.get(str(sid), "")
    cl = "; ".join([f"{curriculum_by_id.get(lk['curriculum_id'], {}).get('lesson_topic', '?')} ({lk['relevance']})"
                    for lk in song_curriculum.get(sid, [])[:5]])
    el = "; ".join([f"{eyt_by_id.get(lk['early_years_id'], {}).get('lesson_goal', '?')} ({lk['relevance']})"
                    for lk in song_early_years.get(sid, [])[:5]])
    write_row(ws1, row, [
        s.get("song_name", ""), s.get("cd_title", "") or "Kathy Reid Naiman",
        s.get("sheet_name", "") or "All Ages",
        infer_type(s.get("song_name"), s.get("actions"), s.get("topic")),
        s.get("topic", "") or "", tags_str, s.get("lyrics", "") or "",
        s.get("actions", "") or "", infer_materials(s.get("actions")),
        s.get("artist", "") or "", s.get("cd_title", "") or "", cl, el
    ], font=TOPIC_FONT)
    row += 1

for sid in sorted(ct_songs.keys()):
    s = ct_songs[sid]
    cl = "; ".join([f"{curriculum_by_id.get(lk['curriculum_id'], {}).get('lesson_topic', '?')} ({lk['relevance']})"
                    for lk in ct_curriculum.get(sid, [])[:5]])
    el = "; ".join([f"{eyt_by_id.get(lk['early_years_id'], {}).get('lesson_goal', '?')} ({lk['relevance']})"
                    for lk in ct_early_years.get(sid, [])[:5]])
    write_row(ws1, row, [
        s.get("song_name", ""), s.get("source", "") or "Circle Time",
        s.get("age_group", "") or "",
        infer_type(s.get("song_name"), s.get("actions"), s.get("category")),
        s.get("teaches", "") or s.get("category", "") or "", "",
        "", s.get("actions", "") or "", infer_materials(s.get("actions")),
        "", s.get("source", "") or "", cl, el
    ], font=TOPIC_FONT)
    row += 1
print(f"Sheet 1: Master Song List — {row - 2} songs")


# ============================================================
# SHEETS 2-4: EARLY YEARS CURRICULUM
# ============================================================
def build_ey_sheet(wb, name, topic_ids, eyt_by_id, ey_to_songs):
    ws = wb.create_sheet(name)
    h = ["Seq", "Category", "Subcategory", "Developmental Goal", "ELOF Ref", "HDLH Lens", "Song Links"]
    write_header(ws, 1, h)
    set_widths(ws, [6, 28, 30, 55, 14, 18, 55])

    # Build ordered groups: category → subcategory → items
    groups = OrderedDict()
    for tid in topic_ids:
        t = eyt_by_id.get(tid)
        if not t:
            continue
        cat = t.get("_category", "(Uncategorized)")
        sub = t.get("_subcategory", "")
        key = (cat, sub)
        if key not in groups:
            groups[key] = []
        groups[key].append(t)

    rn = 2
    prev_cat = None
    for (cat, sub), items in groups.items():
        # Category header (only when category changes)
        if cat != prev_cat:
            write_row(ws, rn, [cat, "", "", "", "", "", ""], font=CAT_FONT, fill=CAT_FILL)
            for c in range(1, 8):
                ws.cell(row=rn, column=c).fill = CAT_FILL
                ws.cell(row=rn, column=c).font = CAT_FONT
            ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=7)
            rn += 1
            prev_cat = cat

        # Subcategory sub-header
        if sub:
            write_row(ws, rn, ["", sub, "", "", "", "", ""], font=SUB_FONT, fill=SUB_FILL)
            for c in range(1, 8):
                ws.cell(row=rn, column=c).fill = SUB_FILL
                ws.cell(row=rn, column=c).font = SUB_FONT
            ws.merge_cells(start_row=rn, start_column=2, end_row=rn, end_column=7)
            rn += 1

        # Data rows
        for t in sorted(items, key=lambda x: int(x.get("seq_num") or 999)):
            sn = ey_to_songs.get(t["id"], [])
            sn_str = "; ".join(sn[:6]) if sn else ""
            write_row(ws, rn, [
                t.get("seq_num", ""), "", "",
                t.get("lesson_goal", ""), t.get("elof_ref", ""),
                t.get("hdldh_lens", ""), sn_str
            ], font=TOPIC_FONT)
            rn += 1

    print(f"  {name}: {rn - 2} rows, {len(groups)} groups")
    return ws


# Daycare (0-2): first 7 seq items per category (infant/toddler level)
daycare_ids = [t["id"] for t in eyt_all
               if t.get("sheet_name") == "Daycare & Preschool"
               and t.get("seq_num") and t["_category"] != "How this sheet works:"
               and t.get("seq_num", "").isdigit() and int(t["seq_num"]) <= 7]

# Preschool (3-4): seq 8+ items (preschool level)
preschool_ids = [t["id"] for t in eyt_all
                 if t.get("sheet_name") == "Daycare & Preschool"
                 and t.get("seq_num") and t["_category"] != "How this sheet works:"
                 and t.get("seq_num", "").isdigit() and int(t["seq_num"]) > 7]

# Kindergarten (5): Music & Arts topics (empty sheet_name) + high-level preschool items
kinder_ids = [t["id"] for t in eyt_all
              if (t.get("sheet_name") in ("", None) and t.get("seq_num"))
              or (t.get("sheet_name") == "Daycare & Preschool"
                  and t.get("seq_num", "").isdigit() and int(t["seq_num"]) > 10)]

build_ey_sheet(wb, "Daycare (0-2)", daycare_ids, eyt_by_id, ey_to_songs)
build_ey_sheet(wb, "Preschool (3-4)", preschool_ids, eyt_by_id, ey_to_songs)
build_ey_sheet(wb, "Kindergarten (5)", kinder_ids, eyt_by_id, ey_to_songs)


# ============================================================
# SHEETS 5-6: GRADE CURRICULUM
# ============================================================
def build_grade_sheet(wb, name, primary, combined, code_prefix):
    ws = wb.create_sheet(name)
    h = ["Subject", "Category", "Seq", "Lesson Topic", "Skill Statement",
         "US Code", "Ontario Code", "Teaching Source", "Song Links", "Teaching Notes"]
    write_header(ws, 1, h)
    set_widths(ws, [22, 30, 6, 42, 45, 16, 20, 25, 55, 35])

    # Tag combined topics
    all_topics = []
    for t in primary:
        all_topics.append({**t, "_src": "primary"})
    for t in combined:
        all_topics.append({**t, "_src": "combined"})

    # Group: subject → category → topics (OrderedDict preserves order)
    subjects = OrderedDict()
    for t in all_topics:
        subj = t.get("subject") or "(No Subject)"
        cat = t.get("category") or "(No Category)"
        if subj not in subjects:
            subjects[subj] = OrderedDict()
        if cat not in subjects[subj]:
            subjects[subj][cat] = []
        subjects[subj][cat].append(t)

    rn = 2
    for subj, cats in subjects.items():
        # Subject header
        write_row(ws, rn, [subj] + [""] * 9, font=CAT_FONT, fill=CAT_FILL)
        for c in range(1, 11):
            ws.cell(row=rn, column=c).fill = CAT_FILL
            ws.cell(row=rn, column=c).font = CAT_FONT
        ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=10)
        rn += 1

        for cat, topics in cats.items():
            # Category sub-header
            write_row(ws, rn, ["", cat] + [""] * 8, font=SUB_FONT, fill=SUB_FILL)
            for c in range(1, 11):
                ws.cell(row=rn, column=c).fill = SUB_FILL
                ws.cell(row=rn, column=c).font = SUB_FONT
            ws.merge_cells(start_row=rn, start_column=2, end_row=rn, end_column=10)
            rn += 1

            for t in sorted(topics, key=lambda x: float(x.get("seq_num") or 999)):
                us = t.get("us_code", "")
                on = t.get("ontario_code", "")
                # For combined topics, tag with grade prefix
                src_label = " ★" if t["_src"] == "combined" else ""
                sn = curr_to_songs.get(t["id"], [])
                sn_str = "; ".join(sn[:6]) if sn else ""
                write_row(ws, rn, [
                    "", "", t.get("seq_num", ""),
                    (t.get("lesson_topic", "") or "") + src_label,
                    t.get("skill_statement", ""),
                    us, on, t.get("teaching_source", ""),
                    sn_str, t.get("teaching_notes", "")
                ], font=TOPIC_FONT)
                rn += 1

    print(f"  {name}: {rn - 2} rows ({len(primary)} primary + {len(combined)} combined)")
    return ws

build_grade_sheet(wb, "Grade 1", g1_topics, g12_topics, "1")
build_grade_sheet(wb, "Grade 2", g2_topics, g12_topics, "2")


# ============================================================
# SAVE
# ============================================================
wb.save(OUTPUT)
print(f"\n✅ Saved: {OUTPUT}")
print(f"   Sheets: {wb.sheetnames}")
